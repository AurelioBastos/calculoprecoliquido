"""
NF-e PreÃ§o LÃ­quido â€” Backend FastAPI
Serve o frontend estÃ¡tico + processa XMLs em memÃ³ria (sem persistÃªncia).
"""
import os, re, json
import unicodedata
from io import BytesIO
from typing import Any

import pandas as pd
import xml.etree.ElementTree as ET

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

app = FastAPI(title="NF-e PreÃ§o LÃ­quido")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# â”€â”€ Serve o frontend â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
HTML_FILE = os.path.join(os.path.dirname(__file__), "nfe_app.html")

@app.get("/", response_class=HTMLResponse)
async def root():
    with open(HTML_FILE, "r", encoding="utf-8") as f:
        return f.read()

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HELPERS â€” PARSER XML
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
NS = re.compile(r'\{[^}]+\}')

def strip_ns(tag: str) -> str:
    return NS.sub('', tag)

def find_text(node, *paths) -> str:
    for path in paths:
        parts = path.split('/')
        cur = [node]
        for part in parts:
            cur = [c for n in cur for c in n if strip_ns(c.tag) == part]
        if cur:
            return (cur[0].text or '').strip()
    return ''

def find_all(node, tag: str):
    return [c for c in node.iter() if strip_ns(c.tag) == tag]

def to_float(s: str) -> float:
    if not s:
        return 0.0
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return 0.0

def parse_date(s: str) -> str:
    return s[:10] if s and len(s) >= 10 else ''

def parse_nfe(file_bytes: bytes, filename: str) -> list[dict]:
    rows = []
    try:
        tree = ET.fromstring(file_bytes)
    except ET.ParseError:
        return rows

    ide   = find_all(tree, 'ide')
    dest  = find_all(tree, 'dest')
    emit  = find_all(tree, 'emit')
    prot  = find_all(tree, 'infProt')
    infAd = find_all(tree, 'infAdic')

    nNF      = find_text(ide[0],  'nNF')   if ide   else ''
    serie    = find_text(ide[0],  'serie') if ide   else ''
    dhEmi    = parse_date(find_text(ide[0], 'dhEmi')) if ide else ''
    CNPJdest = find_text(dest[0], 'CNPJ')  if dest  else ''
    CNPJemit = find_text(emit[0], 'CNPJ')  if emit  else ''
    nProt    = find_text(prot[0], 'nProt') if prot  else ''
    infCpl   = find_text(infAd[0],'infCpl')if infAd else ''

    chNFe = ''
    for el in tree.iter():
        if strip_ns(el.tag) == 'infNFe':
            chNFe = el.get('Id', '').replace('NFe', '')
            break

    for det in find_all(tree, 'det'):
        prod = find_all(det, 'prod')
        imp  = find_all(det, 'imposto')
        p    = prod[0] if prod else det
        nItem = det.get('nItem', '') or ''

        cProd  = find_text(p, 'cProd')
        xProd  = find_text(p, 'xProd')
        NCM    = find_text(p, 'NCM')
        CFOP   = find_text(p, 'CFOP')
        uCom   = find_text(p, 'uCom')
        qCom   = to_float(find_text(p, 'qCom'))
        vUnCom = to_float(find_text(p, 'vUnCom'))
        vProd  = to_float(find_text(p, 'vProd'))
        xPed   = find_text(p, 'xPed')
        nItemPed_s = find_text(p, 'nItemPed')

        if xPed and nItemPed_s:
            try:
                nItemPed = int(float(nItemPed_s))
            except (TypeError, ValueError):
                nItemPed = 0
            xPednItem = f"{xPed}-{nItemPed}" if nItemPed else '0'
        else:
            xPed, nItemPed, xPednItem = '0', 0, '0'

        orig = cst = csticms = ''
        pICMS = bcICMS = vICMS = 0.0
        vBCST = pICMSST = vICMSST = 0.0
        vBCFCPST = pFCPST = vFCPST = pRedBC = 0.0
        bcIPI = pIPI = vIPI = 0.0

        if imp:
            icms_nodes = find_all(imp[0], 'ICMS')
            if icms_nodes:
                for child in icms_nodes[0]:
                    orig     = find_text(child, 'orig')    or orig
                    cst_v    = find_text(child, 'CST')
                    if cst_v: cst = cst_v.zfill(2)
                    pICMS    = to_float(find_text(child, 'pICMS'))    or pICMS
                    bcICMS   = to_float(find_text(child, 'vBC'))      or bcICMS
                    vICMS    = to_float(find_text(child, 'vICMS'))    or vICMS
                    vBCST    = to_float(find_text(child, 'vBCST'))    or vBCST
                    pICMSST  = to_float(find_text(child, 'pICMSST'))  or pICMSST
                    vICMSST  = to_float(find_text(child, 'vICMSST'))  or vICMSST
                    vBCFCPST = to_float(find_text(child, 'vBCFCPST')) or vBCFCPST
                    pFCPST   = to_float(find_text(child, 'pFCPST'))   or pFCPST
                    vFCPST   = to_float(find_text(child, 'vFCPST'))   or vFCPST
                    pRedBC   = to_float(find_text(child, 'pRedBC'))   or pRedBC

            csticms = (orig + cst) if (orig and cst) else 'Simples'
            if not orig: orig = 'Simples'
            if not cst:  cst  = 'Simples'

            ipi_nodes = find_all(imp[0], 'IPI')
            if ipi_nodes:
                for child in ipi_nodes[0]:
                    bcIPI = to_float(find_text(child, 'vBC'))  or bcIPI
                    pIPI  = to_float(find_text(child, 'pIPI')) or pIPI
                    vIPI  = to_float(find_text(child, 'vIPI')) or vIPI

        rows.append({
            '_id': None,
            'CNPJ Emit': CNPJemit, 'CNPJ Dest': CNPJdest,
            'Data EmissÃ£o': dhEmi, 'NÂº NF': nNF, 'SÃ©rie': serie,
            'Chave NF-e': chNFe, 'nProt': nProt,
            'CÃ³d Produto': cProd, 'DescriÃ§Ã£o': xProd,
            'Ped-Item': xPednItem, 'Pedido': xPed, 'Item Ped': nItemPed,
            'nItem': nItem,
            'NCM': NCM, 'Qtd': qCom, 'UN': uCom,
            'Vl Unit': vUnCom, 'Vl Produto': vProd, 'CFOP': CFOP,
            'CST ICMS': csticms, 'Orig': orig, 'CST': cst,
            '% ICMS': pICMS, 'BC ICMS': bcICMS, 'Vl ICMS': vICMS,
            '% IPI': pIPI, 'BC IPI': bcIPI, 'Vl IPI': vIPI,
            '% ICMS-ST': pICMSST, 'BC ICMS-ST': vBCST, 'Vl ICMS-ST': vICMSST,
            '% FCP-ST': pFCPST, 'BC FCP-ST': vBCFCPST, 'Vl FCP-ST': vFCPST,
            '% Red BC': pRedBC, 'Inf Adicionais': infCpl,
            # campos editÃ¡veis (defaults)
            'Fator Conv.': 1.0, 'Multiplicador': 1.0,
            '% PIS+COFINS': None, 'Taxa CÃ¢mbio': None, 'Tipo Material': None,
            # campos calculados (preenchidos pelo recalc)
            'Vl Unit BRL': 0.0, 'Vl Unit Pedido': 0.0, 'Qtd Pedido': 0.0,
            'Vl PIS+COFINS': 0.0, 'PreÃ§o LÃ­q PC': 0.0, 'PreÃ§o LÃ­q Total': 0.0,
        })
    return rows


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# CÃLCULO PREÃ‡O LÃQUIDO
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def calcular_linha(row: dict, pis_rate_global: float, taxa_global: float, tipo_global: str) -> dict:
    fator = float(row.get('Fator Conv.') or 1.0) or 1.0
    mult  = float(row.get('Multiplicador') or 1.0) or 1.0
    tipo  = row.get('Tipo Material') or tipo_global
    q_nfe = float(row.get('Qtd') or 1.0) or 1.0

    # Taxa individual sobrepÃµe global se preenchida
    taxa_row = row.get('Taxa CÃ¢mbio')
    taxa = float(taxa_row) if taxa_row is not None else taxa_global

    # PIS individual sobrepÃµe global se preenchida e > 0
    pis_row = row.get('% PIS+COFINS')
    if pis_row is not None and float(pis_row) > 0:
        pis_rate = float(pis_row) / 100.0
    else:
        pis_rate = pis_rate_global

    vUnit      = float(row.get('Vl Unit') or 0)
    vUnit_ped  = vUnit / fator
    qtd_pedido = q_nfe * fator

    def norm(v):
        v = float(v) if v else 0.0
        return v / 100.0 if v > 1.0 else v

    pICMS  = norm(row.get('% ICMS'))
    pIPI   = norm(row.get('% IPI'))
    pRedBC = norm(row.get('% Red BC'))
    pICMS_ef = pICMS * (1 - pRedBC)

    if tipo == 'Ativo/Consumo':
        BC          = vUnit_ped * (1 + pIPI)
        vIPI_ped    = vUnit_ped * pIPI
        vICMS_ped   = BC * pICMS_ef
        vPisCofins  = BC * pis_rate
        conversao_total = BC - vICMS_ped - vPisCofins - vIPI_ped
    else:
        BC          = vUnit_ped
        vIPI_ped    = 0.0
        vICMS_ped   = BC * pICMS_ef
        vPisCofins  = BC * pis_rate
        conversao_total = BC - vICMS_ped - vPisCofins

    preco_liq   = conversao_total / taxa if taxa != 0 else conversao_total
    vUnit_brl   = vUnit / taxa if taxa != 0 else vUnit
    preco_total = preco_liq * mult

    return {
        'Qtd Pedido':     round(qtd_pedido, 2),
        'Vl Unit BRL':    round(vUnit_brl,  2),
        'Vl Unit Pedido': round(vUnit_ped,  2),
        'Vl PIS+COFINS':  round(vPisCofins, 2),
        'PreÃ§o LÃ­q PC':   round(preco_liq,  2),
        'PreÃ§o LÃ­q Total':round(preco_total,2),
    }


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ROTAS API
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.post("/api/upload-xml")
async def upload_xml(files: list[UploadFile] = File(...)):
    all_rows = []
    for f in files:
        content = await f.read()
        all_rows.extend(parse_nfe(content, f.filename))
    if not all_rows:
        raise HTTPException(400, "Nenhum item encontrado nos XMLs enviados.")
    for idx, row in enumerate(all_rows, start=1):
        row['_id'] = idx
    return {"data": all_rows}


class RecalcPayload(BaseModel):
    data: list[dict]
    pis_rate: float = 0.0
    taxa_efetiva: float = 1.0
    tipo_global: str = "Ativo/Consumo"

@app.post("/api/recalc")
async def recalc(payload: RecalcPayload):
    pis_rate = payload.pis_rate / 100.0
    for row in payload.data:
        calc = calcular_linha(row, pis_rate, payload.taxa_efetiva, payload.tipo_global)
        row.update(calc)
    return {"data": payload.data}


@app.post("/api/procv-preview")
async def procv_preview(ref_file: UploadFile = File(...)):
    content = await ref_file.read()
    fname = ref_file.filename.lower()
    try:
        if fname.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            df = pd.read_excel(BytesIO(content), header=0)
        df.columns = df.columns.astype(str).str.strip()
        cols = list(df.columns)
        preview = df.head(5).fillna('').astype(str).to_dict(orient='records')
        return {"columns": cols, "preview": preview}
    except Exception as e:
        raise HTTPException(400, str(e))


@app.post("/api/procv-apply")
async def procv_apply(
    ref_file:   UploadFile = File(...),
    data:       str        = Form(...),
    col_chave:  str        = Form(...),
    col_chave_2:str        = Form(''),
    col_pedido: str        = Form(...),
    col_item:   str        = Form(...),
    campo_xml:  str        = Form('nItem'),
    campo_xml_2:str        = Form(''),
):
    content = await ref_file.read()
    fname = ref_file.filename.lower()
    try:
        if fname.endswith('.csv'):
            df_ref = pd.read_csv(BytesIO(content))
        else:
            df_ref = pd.read_excel(BytesIO(content), header=0)
        df_ref.columns = df_ref.columns.astype(str).str.strip()
    except Exception as e:
        raise HTTPException(400, str(e))

    rows = json.loads(data)

    def _norm(v):
        s = str(v).strip()
        if s.lower() in ('nan','none',''): return ''
        try:
            f = float(s)
            return str(int(f)) if f == int(f) else s
        except ValueError:
            return s

    # Monta lookup: chave1 (+ chave2 opcional) â†’ "PEDIDO-ITEM"
    lookup: dict[tuple, str] = {}
    for _, row in df_ref.iterrows():
        k1 = _norm(row[col_chave])
        k2 = _norm(row[col_chave_2]) if col_chave_2 and col_chave_2 in df_ref.columns else ''
        ped  = _norm(row[col_pedido]).lstrip('0') or '0'
        item = _norm(row[col_item]).lstrip('0')   or '0'
        if k1:
            lookup[(k1, k2)] = f"{ped}-{item}"

    campo_map = {'nItem':'nItem','CÃ³d Produto':'CÃ³d Produto','DescriÃ§Ã£o':'DescriÃ§Ã£o','Item Ped':'Item Ped'}
    campo_map2= {'nItem':'nItem','CÃ³d Produto':'CÃ³d Produto','DescriÃ§Ã£o':'DescriÃ§Ã£o','Item Ped':'Item Ped'}

    encontrados = nao_encontrados = 0
    for row in rows:
        k1 = _norm(row.get(campo_map.get(campo_xml, campo_xml), ''))
        k2 = _norm(row.get(campo_map2.get(campo_xml_2,''), '')) if campo_xml_2 else ''
        match = lookup.get((k1, k2)) or lookup.get((k1, ''))
        if match:
            row['Ped-Item'] = match
            encontrados += 1
        else:
            nao_encontrados += 1

    return {"data": rows, "encontrados": encontrados, "nao_encontrados": nao_encontrados}


@app.post("/api/confronto-pc")
async def confronto_pc(
    pc_file: UploadFile = File(...),
    data:    str        = Form(...),
):
    content = await pc_file.read()
    def _norm_col(v: Any) -> str:
        s = str(v or "").strip()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        s = s.lower()
        return re.sub(r"[^a-z0-9]+", "", s)

    aliases = {
        "documento": {"documento", "doc", "numerodocumento", "nrdocumento"},
        "item": {"item", "it"},
        "ped_item": {"peditem", "pedidoitem", "pedidoitemchave", "chave", "chavepc"},
        "vl_liq_unit": {
            "vlliqunit", "vlliqunit", "viliqunit", "viliqunit",
            "vlliquidounit", "vlliqunitario", "valorliquidounitario"
        },
        "aliq_icms": {"aliqicms", "aliqicm", "icms"},
        "aliq_ipi": {"aliqipi", "ipi"},
        "aliq_st_icms": {"aliqsticms", "sticms", "aliqst"},
        "ncm": {"ncm"},
        "origem": {"origem", "orig"},
    }

    # Documento+Item OU Ped-Item, além das demais colunas fiscais
    required_keys = ["vl_liq_unit", "aliq_icms", "aliq_ipi", "aliq_st_icms", "ncm", "origem"]

    def _resolve_cols(df: pd.DataFrame) -> dict[str, str]:
        norm_to_raw: dict[str, str] = {}
        for c in df.columns:
            nc = _norm_col(c)
            if nc and nc not in norm_to_raw:
                norm_to_raw[nc] = c

        resolved: dict[str, str] = {}
        for key, names in aliases.items():
            hit = next((raw for n, raw in norm_to_raw.items() if n in names), None)
            if hit:
                resolved[key] = hit

        # Fallback heuristico para variacoes comuns de coluna (ex.: VI.Liq.Unit.)
        def _find_contains(*parts: str):
            for n, raw in norm_to_raw.items():
                if all(part in n for part in parts):
                    return raw
            return None

        if "vl_liq_unit" not in resolved:
            h = _find_contains("liq", "unit")
            if h:
                resolved["vl_liq_unit"] = h
        if "aliq_icms" not in resolved:
            h = _find_contains("aliq", "icms")
            if h:
                resolved["aliq_icms"] = h
        if "aliq_ipi" not in resolved:
            h = _find_contains("aliq", "ipi")
            if h:
                resolved["aliq_ipi"] = h
        if "aliq_st_icms" not in resolved:
            h = _find_contains("aliq", "st", "icms")
            if h:
                resolved["aliq_st_icms"] = h
        return resolved

    # Tenta diferentes abas e linhas de cabecalho
    # (planilhas com capa/legenda antes da tabela real)
    best_df = None
    best_map: dict[str, str] = {}
    best_score = -1
    best_sheet = None
    best_header = None
    last_err = None
    try:
        xls = pd.ExcelFile(BytesIO(content))
        sheet_names = xls.sheet_names or [0]
    except Exception:
        sheet_names = [0]

    for sheet in sheet_names:
        for header_row in range(0, 31):
            try:
                probe = pd.read_excel(BytesIO(content), sheet_name=sheet, header=header_row)
                probe.columns = probe.columns.astype(str).str.strip()
                cmap = _resolve_cols(probe)
                score = len(cmap)
                # Bonus leve para tabelas com mais linhas validas
                if score == best_score and best_df is not None:
                    cur_rows = len(probe.dropna(how="all"))
                    best_rows = len(best_df.dropna(how="all"))
                    if cur_rows <= best_rows:
                        continue
                if score > best_score or (score == best_score and best_df is not None and len(probe.dropna(how="all")) > len(best_df.dropna(how="all"))):
                    best_score = score
                    best_df = probe
                    best_map = cmap
                    best_sheet = sheet
                    best_header = header_row
            except Exception as e:
                last_err = e

    if best_df is None:
        raise HTTPException(400, str(last_err) if last_err else "Falha ao ler planilha de PC.")

    missing_keys = [k for k in required_keys if k not in best_map]
    if missing_keys:
        disp = list(best_df.columns)
        raise HTTPException(
            400,
            f"Colunas nao reconhecidas no PC: {missing_keys}. "
            f"Aba/linha testada com melhor resultado: {best_sheet}/{best_header}. "
            f"Disponiveis: {disp}"
        )
    has_doc_item = ("documento" in best_map and "item" in best_map)
    has_ped_item = ("ped_item" in best_map)
    if not has_doc_item and not has_ped_item:
        disp = list(best_df.columns)
        raise HTTPException(
            400,
            f"Nao encontrei colunas de chave do PC (Documento+Item ou Ped-Item). "
            f"Aba/linha testada com melhor resultado: {best_sheet}/{best_header}. "
            f"Disponiveis: {disp}"
        )

    df_pc = best_df

    def _norm_ped_item(v: Any) -> str:
        s = str(v or "").strip()
        if not s:
            return ""
        s = s.replace(" ", "")
        if "-" in s:
            a, b = s.split("-", 1)
            return f"{a}-{b.lstrip('0') or '0'}"
        return s

    if has_doc_item:
        df_pc['Chave PC'] = (
            df_pc[best_map["documento"]].astype(str).str.strip() + '-' +
            df_pc[best_map["item"]].astype(str).str.strip().str.lstrip('0')
        )
    else:
        df_pc['Chave PC'] = df_pc[best_map["ped_item"]].map(_norm_ped_item)
    df_pc_key = df_pc.drop_duplicates(subset='Chave PC').set_index('Chave PC')

    rows = json.loads(data)
    result = []

    def safe_pct(v):
        try:
            f = float(v) if v is not None else 0.0
            return round(f * 100, 4) if f <= 1.0 else round(f, 4)
        except: return 0.0

    div_vl = div_icms = div_ipi = div_st = div_ncm = div_orig = sem_match = matches = 0

    for row in rows:
        chave = str(row.get('Ped-Item', '')).strip()
        base = {
            '_id': row.get('_id'),
            'DescriÃ§Ã£o':  row.get('DescriÃ§Ã£o',''),
            'Ped-Item':   chave,
            'ICMS XML (%)': safe_pct(row.get('% ICMS')),
            'IPI XML (%)':  safe_pct(row.get('% IPI')),
            'ICMS-ST XML (%)': safe_pct(row.get('% ICMS-ST')),
            'NCM XML':    str(row.get('NCM','')).strip().replace('.',''),
            'Origem XML': str(row.get('Orig','')).strip(),
            'PreÃ§o LÃ­q Total (XML)': row.get('PreÃ§o LÃ­q Total', 0),
        }

        if chave in df_pc_key.index:
            matches += 1
            pc = df_pc_key.loc[chave]
            vl_xml = float(row.get('PreÃ§o LÃ­q Total') or 0)
            pc_vl = pc[best_map["vl_liq_unit"]]
            vl_pc  = float(str(pc_vl).replace(',','.')) if pd.notna(pc_vl) else 0.0
            dif_vl = round(vl_xml - vl_pc, 2)

            lim_tol = abs(vl_pc) * 0.15
            dentro  = abs(dif_vl) <= lim_tol
            if abs(dif_vl) <= 0.001: st_dif = 'OK âœ…'
            elif dentro:             st_dif = 'TOL âœ…'
            else:                    st_dif = 'DIVERGENTE âš ï¸'; div_vl += 1

            def cmp(xml_val, col):
                xp = safe_pct(xml_val)
                vpc = pc[col]
                try: pp = round(float(str(vpc).replace(',','.')), 4) if pd.notna(vpc) else 0.0
                except: pp = 0.0
                return ('OK âœ…' if abs(xp-pp)<0.0001 else 'DIVERGENTE âš ï¸'), xp, pp

            st_icms, xi, pi_ = cmp(row.get('% ICMS'),    best_map["aliq_icms"])
            st_ipi,  xi2,pi2 = cmp(row.get('% IPI'),     best_map["aliq_ipi"])
            st_st,   xs, ps  = cmp(row.get('% ICMS-ST'), best_map["aliq_st_icms"])

            ncm_xml = str(row.get('NCM','')).strip().replace('.','')
            pc_ncm = pc[best_map["ncm"]]
            ncm_pc  = str(pc_ncm).strip().replace('.','') if pd.notna(pc_ncm) else ''
            st_ncm  = 'OK âœ…' if ncm_xml == ncm_pc else 'DIVERGENTE âš ï¸'

            orig_xml = str(row.get('Orig','')).strip()
            pc_orig = pc[best_map["origem"]]
            orig_pc  = str(pc_orig).strip() if pd.notna(pc_orig) else ''
            st_orig  = 'OK âœ…' if orig_xml == orig_pc else 'DIVERGENTE âš ï¸'

            if st_icms != 'OK âœ…': div_icms += 1
            if st_ipi  != 'OK âœ…': div_ipi  += 1
            if st_st   != 'OK âœ…': div_st   += 1
            if st_ncm  != 'OK âœ…': div_ncm  += 1
            if st_orig != 'OK âœ…': div_orig  += 1

            result.append({**base,
                'Vl LÃ­q Unit PC': round(vl_pc, 2), 'Dif. Vl Unit': dif_vl,
                'Lim. TolerÃ¢ncia': round(lim_tol,2), 'Status Dif.': st_dif,
                'ICMS PC (%)': pi_, 'Status ICMS': st_icms,
                'IPI PC (%)':  pi2, 'Status IPI':  st_ipi,
                'ICMS-ST PC (%)': ps, 'Status ICMS-ST': st_st,
                'NCM PC': ncm_pc, 'Status NCM': st_ncm,
                'Origem PC': orig_pc, 'Status Origem': st_orig,
                'Encontrado': True,
            })
        else:
            sem_match += 1
            result.append({**base,
                'Vl LÃ­q Unit PC': None, 'Dif. Vl Unit': None,
                'Lim. TolerÃ¢ncia': None, 'Status Dif.': 'SEM MATCH âŒ',
                'ICMS PC (%)': None, 'Status ICMS': 'SEM MATCH âŒ',
                'IPI PC (%)':  None, 'Status IPI':  'SEM MATCH âŒ',
                'ICMS-ST PC (%)': None, 'Status ICMS-ST': 'SEM MATCH âŒ',
                'NCM PC': None, 'Status NCM': 'SEM MATCH âŒ',
                'Origem PC': None, 'Status Origem': 'SEM MATCH âŒ',
                'Encontrado': False,
            })

    kpis = dict(
        matches=matches, sem_match=sem_match,
        div_vl=div_vl, div_icms=div_icms, div_ipi=div_ipi, div_st=div_st,
        div_ncm=div_ncm, div_orig=div_orig,
    )
    return {"data": result, "kpis": kpis}


# â”€â”€ Helpers de estilo Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_C_HDR_BG  = "0B2F1F"; _C_HDR_FG  = "ECFFF2"
_C_INFO_BG = "102A1D"; _C_INFO_FG = "B0D7BC"
_C_ODD     = "0A1F16"; _C_EVEN    = "0F3626"
_C_OK      = "2CCF66"; _C_WARN    = "F2B84B"
_C_BAD     = "F66A6A"; _C_NM      = "77A88B"
_C_MONEY   = "ECFFF2"; _C_BORDER  = "1F4A33"

_thin   = Side(style="thin", color=_C_BORDER)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _style(cell, bg, fg, bold=False, align="left", num_fmt=None):
    cell.font      = Font(bold=bold, color=fg, name="Arial", size=9)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border
    if num_fmt: cell.number_format = num_fmt
def _status_fg(val):
    s = str(val or "")
    if "DIVERGENTE" in s: return _C_BAD
    if "TOL"        in s: return _C_WARN
    if "OK"         in s: return _C_OK
    return _C_NM
def _auto_width(ws, mn=8, mx=42):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2,mn),mx)
def _clean(val):
    if isinstance(val, str):
        return val.replace("âœ…","").replace("âš ï¸","").replace("âŒ","").strip()
    return val

_STATUS_COLS = {"Status Dif.","Status ICMS","Status IPI","Status NCM","Status Origem","Status ICMS-ST","Encontrado"}
_MONEY_COLS  = {"PreÃ§o LÃ­q Total (XML)","Vl LÃ­q Unit PC","Dif. Vl Unit","Lim. TolerÃ¢ncia",
                "PreÃ§o LÃ­q Total","PreÃ§o LÃ­q PC","Vl Unit BRL","Vl Unit Pedido","Vl PIS+COFINS",
                "Vl Produto","Vl Unit","BC ICMS","Vl ICMS","BC IPI","Vl IPI","BC ICMS-ST","Vl ICMS-ST"}
_PCT_COLS    = {"% ICMS","% IPI","% ICMS-ST","% FCP-ST","% Red BC","% PIS+COFINS",
                "ICMS XML (%)","ICMS PC (%)","IPI XML (%)","IPI PC (%)","ICMS-ST XML (%)","ICMS-ST PC (%)"}

def _write_row(ws, excel_row, cols, row_data, is_header=False, bg=None):
    for ci, col in enumerate(cols, 1):
        val = _clean(row_data.get(col) if isinstance(row_data, dict) else row_data[ci-1])
        if col in _STATUS_COLS: fg, bold = _status_fg(val), True
        elif col in _MONEY_COLS: fg, bold = _C_MONEY, False
        else: fg, bold = (_C_HDR_FG, True) if is_header else (_C_HDR_FG, False)
        num_fmt = None
        if col in _MONEY_COLS and val is not None:
            try: val = float(val); num_fmt = '#,##0.00'
            except: pass
        align = "center" if is_header or col in _STATUS_COLS else ("right" if col in (_MONEY_COLS|_PCT_COLS) else "left")
        c = ws.cell(row=excel_row, column=ci, value=val)
        _style(c, bg or _C_HDR_BG, fg, bold=bold, align=align, num_fmt=num_fmt)
    ws.row_dimensions[excel_row].height = 20 if is_header else 16


class ExportPayload(BaseModel):
    data: list[dict]
    pis_rate: float = 0.0
    taxa_efetiva: float = 1.0
    tipo_global: str = "Ativo/Consumo"

@app.post("/api/export-excel")
async def export_excel(payload: ExportPayload):
    from openpyxl import Workbook
    rows = payload.data
    if not rows: raise HTTPException(400, "Sem dados.")
    cols = [c for c in rows[0].keys() if c != "_id"]
    wb = Workbook(); ws = wb.active
    ws.title = "NF-e ICMS"; ws.sheet_view.showGridLines = False
    info = (f"Tipo: {payload.tipo_global}   |   PIS+COFINS: {payload.pis_rate:.4f}%"
            f"   |   Taxa cÃ¢mbio: {payload.taxa_efetiva:.4f}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    _style(ws.cell(row=1, column=1, value=info), _C_INFO_BG, _C_INFO_FG, bold=True)
    ws.row_dimensions[1].height = 18
    _write_row(ws, 2, cols, {c:c for c in cols}, is_header=True)
    for ri, row in enumerate(rows):
        _write_row(ws, ri+3, cols, row, bg=_C_ODD if ri%2==0 else _C_EVEN)
    _auto_width(ws); ws.freeze_panes = "A3"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_icms.xlsx"})


class ExportConfrontoPayload(BaseModel):
    data: list[dict]

@app.post("/api/export-confronto")
async def export_confronto(payload: ExportConfrontoPayload):
    from openpyxl import Workbook
    rows = payload.data
    if not rows: raise HTTPException(400, "Sem dados.")
    ORDERED = ["DescriÃ§Ã£o","Ped-Item","PreÃ§o LÃ­q Total (XML)","Vl LÃ­q Unit PC","Dif. Vl Unit",
               "Lim. TolerÃ¢ncia","Status Dif.","ICMS XML (%)","ICMS PC (%)","Status ICMS",
               "IPI XML (%)","IPI PC (%)","Status IPI","ICMS-ST XML (%)","ICMS-ST PC (%)",
               "Status ICMS-ST","NCM XML","NCM PC","Status NCM","Origem XML","Origem PC",
               "Status Origem","Encontrado"]
    available = set(rows[0].keys())
    cols = [c for c in ORDERED if c in available]
    cols += [c for c in rows[0].keys() if c not in cols and c != "_id"]
    wb = Workbook(); ws = wb.active
    ws.title = "Confronto PC"; ws.sheet_view.showGridLines = False
    _write_row(ws, 1, cols, {c:c for c in cols}, is_header=True)
    for ri, row in enumerate(rows):
        _write_row(ws, ri+2, cols, row, bg=_C_ODD if ri%2==0 else _C_EVEN)
    _auto_width(ws); ws.freeze_panes = "A2"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=confronto_pc.xlsx"})
